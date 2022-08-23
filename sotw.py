# -*- coding: utf-8 -*-
"""
PROGRAM: sotw.py
USAGE: Calculates and documents scores and winner of Song of the Week (SOTW) competition.
AUTHOR: D. Joseph
CREATED ON: Fri Aug 20 15:59:23 2021
"""

##################
# IMPORT MODULES #
##################

import pandas as pd
from datetime import date
from collections import defaultdict
import itertools
import random
import bisect
from openpyxl import load_workbook
import gspread
import spotipy
from spotipy.oauth2 import SpotifyClientCredentials
import config
from os import path

########################################################
# CONSIDER UPDATING THESE VARIABLES PRIOR TO EACH SOTW #
########################################################

year = 2022 #year of competition
month = 8 #month of competition
day = 12 #day of competition
excel_row_old_num = 203 #last populated row in "All-Time Results" tab of Excel results spreadsheet
sp_uri = config.sp_uri #path to Spotify playlist; contained in config.py file so that any potentially sensitive information is not pushed to Github
offset = 0 #start position of where to read in songs from the playlist, where offset = 0 would read in most recently submitted song, offset = 1 would start read in at song #2 in playlist, etc.; in most cases offset will equal 0
song_count = 9 #number of submitted songs for the week
pl_start_pos = 0 + offset #see comment next to creation of offset variable
pl_end_pos =  pl_start_pos + song_count #end position of where to read in songs from the playlist; any non-continuous subset of songs to be read in would need to be hard-coded
DATA_DIR = config.DATA_DIR_name #filepath of folder containing Excel results spreadsheet; contained in config.py file so that any potentially sensitive information is not pushed to Github

################################
# IMPORT DATA FROM SPOTIFY API #
################################

#set up Spotify API credentials;
#client_id & client_secret contained in config.py file so that any potentially sensitive information is not pushed to Github
sp = spotipy.Spotify(client_credentials_manager=SpotifyClientCredentials(
    client_id=config.client_id,
    client_secret=config.client_secret)) 

#import playlist from Spotify API
playlist = sp.playlist(sp_uri)
results = playlist['tracks']
tracks = results['items']
while results['next']:
    results = sp.next(results)
    tracks.extend(results['items'])

#create dictionary with pertinent data fields from Spotify API    
sotw_sp = defaultdict(list)
for track in tracks:
    track_artists = []
    sotw_sp['track_name'].append((track['track']['name']))
    if len(track['track']['artists']) > 1:
        for artist in track['track']['artists']:
            track_artists.append((artist['name']))
        sotw_sp['track_artists'].append((', '.join(track_artists)))
    else:
        sotw_sp['track_artists'].append((track['track']['artists'][0]['name']))
    sotw_sp['added_at'].append((track['added_at']))
    sotw_sp['added_by'].append((track['added_by']['id']))
    sotw_sp['popularity'].append((track['track']['popularity']))

#create DataFrame from dictionary of full playlist
sotw_sp_df = pd.DataFrame(sotw_sp)
sotw_sp_df = sotw_sp_df.sort_values("added_at", ascending=False)



#DEFAULT IS COMMENTED OUT: hard code in case any song needs to be dropped from songs (i.e. non-continuous subset of submitted songs in playlist)
'''
sotw_sp_df.drop(sotw_sp_df[(sotw_sp_df['track_name'] == '<Song Name>') &
                           (sotw_sp_df['track_artists'] == '<Artist Name>')].index, 
                           inplace=True)
'''



#limit DataFrame to this week's submitted songs
sotw_sp_df_10 = sotw_sp_df.iloc[pl_start_pos:pl_end_pos].reset_index().drop(columns='index')

#dictionary with Spotify usernames and names of usual SOTW participants;
#dictionary contained in config.py file so that any potentially sensitive information is not pushed to Github
usernames = config.usernames_dict

#replace Spotify usernames with names in added_by field in DataFrame
for key,value in usernames.items():
    sotw_sp_df_10.loc[sotw_sp_df_10['added_by']==key,'added_by'] = value

#DEFAULT IS COMMENTED OUT: hard code for cleaning Spotify track_name if necessary



#dictionary of clean song names (keys) and participant names (values) in this week's SOTW; contained in config.py file
clean_songs = config.clean_songs_dict 
for key,value in clean_songs.items():
    sotw_sp_df_10.loc[sotw_sp_df_10['added_by']==value,'track_name'] = key



#create lists for song names, song artists, and SOTW participant names
track_name_this_week = list(sotw_sp_df_10['track_name'])
track_artists_this_week = list(sotw_sp_df_10['track_artists'])
added_by_this_week = list(sotw_sp_df_10['added_by'])

#DEFAULT IS COMMENTED OUT: hard code for adjusting added_by value if song is submitted by someone other than the specified participant

'''
added_by_this_week[0] = '<Participant Name>'
'''


this_week = list(zip(track_name_this_week, track_artists_this_week, added_by_this_week))

#dictionary of songs (keys) and participant names (values) in this week's SOTW
songs_submitter = {}
for x in this_week:
    songs_submitter["{}{} - {}".format(x[0],'"',x[1])] = "{}".format(x[2])
submitter_songs = {v: k for k, v in songs_submitter.items()} #submitter/song dictionary (inverse of above dictionary)
submitters = list(songs_submitter.values()) #list of submitters (dependent on first dictionary)

##########################################
# IMPORT EXCEL & GOOGLE SPREADSHEET DATA #
##########################################

#assign variables needed for import
date_xlsx_cell = date(year, month, day).strftime("%#m/%#d/%Y") #dash-mdy-format of date that will go into Excel cells
date_xlsx_fn = date(year, month, day).strftime("%Y.%m.%d") #slash-Ymd-format of date that will go into Excel filename
date_google = date(year, month, day).strftime("%Y.%m.%d") #slash-Ymd-format of date that comes from Google sheet; typically will correspond w/ date_xlsx_fn
excel_row_old = str(excel_row_old_num) #Excel row where data was written to in last SOTW/last time program was run
excel_row_new = str(excel_row_old_num + 1) #Excel row where data will be written to in this SOTW/program run

#load Excel file
fn = path.join(DATA_DIR, "Song of the Week_{}.xlsx".format(date_xlsx_fn)) 
wb = load_workbook(fn)
ws1 = wb["Points Calculation Sheet"]
ws2 = wb["All-Time Results"]
ws3 = wb["Available Total Points"]

#read in Google spreadsheet
gc = gspread.service_account()
sh = gc.open("{} NERA Weekly Music Competition (Responses)".format(date_google))
sheet1 = sh.sheet1

#create DataFrame from Google spreadsheet
sotw_input = pd.DataFrame(sheet1.get_all_records())
sotw_copy = sotw_input.copy() #created copy of input DataFrame just to memorialize initial input

###############################################################################
# CLEAN GOOGLE SPREADSHEET DATAFRAME AND TURN IT INTO SOTW POINTS CALCULATION #
###############################################################################

#clean song name columns from SOTW Google spreadsheet DataFrame
song_names = []
for i in sotw_copy.columns:
    if i[0]!="W":
        song_names.append(i)
    if i[0]=="W":
        song_names.append(i.split(") [")[1][1:-1])
sotw_sn = sotw_copy.copy()
sotw_sn.columns = song_names

sotw_sn = sotw_sn.rename(columns = songs_submitter) #use songs_submitter dictionary to convert column names to be submitter names rather than song names
#comment to indicate that if need be, this would be a good spot to hard code the fix to reconcile mismatched song names between Spotify and Google Forms
sotw_sn.drop(columns=[col for col in sotw_sn if col not in submitters], inplace=True)  #drop unnecessary columns
sotw_sn = sotw_sn.T #transpose DataFrame so that submitters correspond to rows rather than columns (and votes correspond with columns rather than rows)



# RUN PROGRAM THROUGH HERE TO CHECK IF GOOGLE SPREADSHEET & SPOTIFY SONG NAMES MATCH #                  



#create separate numerical fields for total first, second, third, and points, respectively
sotw_sn["first_place"] = (sotw_sn[list(sotw_sn.columns)] == "First Place").sum(axis=1)
sotw_sn["second_place"] = (sotw_sn[list(sotw_sn.columns)] == "Second Place").sum(axis=1)
sotw_sn["third_place"] = (sotw_sn[list(sotw_sn.columns)] == "Third Place").sum(axis=1)
sotw_sn["points"] = (sotw_sn["first_place"]*3 + sotw_sn["second_place"]*2 + sotw_sn["third_place"])

sotw_points = sotw_sn.sort_values("points", ascending=False)[['first_place', 'second_place', 'third_place', 'points']] #sort by total points and leave only relevant columns
sotw_points.reset_index(inplace=True) #reset index so that it is counter
sotw_points.rename(columns = {"index":"submitters"}, inplace=True) #rename field that was formerly index

#assign maximum points scored, how many winners there are, who winner is if there is only one, and who tiebreak winner is if there are multiple winners
print(sotw_points["points"].sum()) #prints total amount of points in this week's SOTW
max_points = sotw_points["points"].max()
winner_count = (sotw_points["points"]==max_points).sum()
if winner_count == 1:
    winner = sotw_points.loc[0,"submitters"]
    winning_song = submitter_songs[winner].replace('"','')
    print("Winner = " + winner + "\nWinning Song = " + winning_song)
elif winner_count > 1:
    odds = 999/winner_count
    winners = []
    weighted_choices = []
    for i in range(winner_count):
        winners.append(sotw_points.loc[i,"submitters"])
    winners_string = ', '.join(x for x in winners)
    print("Winners = " + winners_string)
    #tiebreaker code below creates weighted probabilities, arranges them in 
    #cumulative distribution using itertools.accumulate(), then locates random
    #value in that cumulative distribution using bisect.bisect()
    for i in winners:
        weighted_choices.append([i, odds])
    weighted_choices.append(['David J.', 1])
    choices, weights = zip(*weighted_choices)
    cumdist = list(itertools.accumulate(weights))
    x = random.random() * cumdist[-1]
    tiebreak_winner = choices[bisect.bisect(cumdist, x)]
    tiebreak_winning_song = submitter_songs[tiebreak_winner]
    print("Tiebreak Winner = " + tiebreak_winner + "\nTiebreak Winning Song = " + tiebreak_winning_song)
sotw_points

#####################################################################
# CREATE TOTAL POINTS ROW TO APPEND TO ALL-TIME RESULTS EXCEL SHEET #
#####################################################################

sotw_points_tp = sotw_points.set_index("submitters").T.drop(["first_place", "second_place", "third_place"]) #transpose total points column into row
sotw_points_tp = sotw_points_tp.reindex(sorted(sotw_points_tp.columns), axis=1) #rearrange columns so that they are in alphabetical order

#####################################################################################
# CREATE AVAILABLE TOTAL POINTS ROW TO APPEND TO AVAILABLE TOTAL POINTS EXCEL SHEET #
#####################################################################################

total_points = int(sotw_points_tp.sum(axis=1)) #sum total points scored by everyone this week

######################################################################
# UPDATE POINTS CALCULATION EXCEL SHEET WITH SOTW POINTS CALCULATION #
######################################################################

#clear points calculation sheet
for row in ws1['A2:E12']:
  for cell in row:
    cell.value = None

sotw_points_array = sotw_points[["submitters", "first_place", "second_place", "third_place", "points"]].to_numpy() #convert sotw_points DataFrame into array
startrow = 2 #set startrow for Excel worksheet iteration
startcol = 1 #set startcolumn for Excel worksheet iteration

#started working on iteration using iterrows(), but didn't pursue; memorializing as reference to possible method in future
'''for rowIndex, row in sotw_points.iterrows1(): #iterate over rows1
    for columnIndex, value in row.items():
        print(value)'''

#update points calculation Excel sheet with SOTW points calculation
for row in sotw_points_array:
    for value in row:
        ws1.cell(startrow, startcol).value = value
        startcol += 1
        if startcol == 6:
            startcol = startcol - 5
            startrow += 1
            continue
        if startrow == 13:
            break

#ws1_updated = pd.DataFrame(ws1.values) #way to convert Excel worksheet into DataFrame

#######################################
# UPDATE ALL-TIME RESULTS EXCEL SHEET #
#######################################

#NEED TO UPDATE THIS TO BE DYNAMIC WHEN WE HAVE LESS THAN FULL SET OF PARTICIPANTS'''

ws2_col1 = ["A{}","B{}","C{}","F{}"]
if winner_count == 1:
    ws2_val1 = [date_xlsx_cell, winner, winning_song, date_xlsx_cell]
elif winner_count > 1:
    ws2_val1 = [date_xlsx_cell, tiebreak_winner, tiebreak_winning_song, date_xlsx_cell]
ws2_col2 = ["H{}","I{}","K{}","L{}","N{}","O{}","R{}","S{}","T{}","U{}"]
ws2_val2 = config.ws2_val2_list #list contained in config.py file so that any potentially sensitive information is not pushed to Github

for col, val in zip(ws2_col1,ws2_val1):
    ws2[col.format(excel_row_new)].value = val


    
#DEFAULT IS COMMENTED OUT: HARD CODE TO TEMPORARILY FIX FOR PREVIOUS THREE SINGLE QUOTE MESSAGE'''
'''
sotw_points_tp["<Participant Name"] = int() #then have to change 0 to missing manually in Excel spreadsheet
'''



for col, val in zip(ws2_col2,ws2_val2):
    ws2[col.format(excel_row_new)].value= int(sotw_points_tp[val])
    
#############################################
# UPDATE AVAILABLE TOTAL POINTS EXCEL SHEET #
#############################################

ws3_col = ["B{}","C{}","D{}","E{}","F{}","G{}","H{}","I{}","J{}","K{}","L{}","M{}","N{}","O{}","P{}"]

ws3["A{}".format(excel_row_new)].value = date_xlsx_cell

for col in ws3_col:
    ws3[col.format(excel_row_new)].value = ws3[col.format(excel_row_old)].value.replace(excel_row_old,excel_row_new)

########################
# SAVE/EXPORT TO EXCEL #
########################

wb.save(fn)
      
#################
# MISCELLANEOUS #
#################
      
#shows 1) how to limit to subset of Excel worksheet, and 2 how to do vlookup in python
'''for row in ws1["A2":"E3"]:
    for cell in row:
        cell.value = "=VLOOKUP(A{}, " '''

#shows how to use .iloc to view subset of DataFrame by column/row numbers
#sotw_points.iloc[0:1, 0:3]

#attempts to pull tables from Excel worksheet; wasn't having luck with this, so have to look into further
'''wb_cc = load_workbook("CC.xlsx")
ws1_cc = wb_cc.worksheets[1]
ws1_cc.tables
ws1_cc_tables = []
for table in ws1_cc.tables:
    ws1_cc_tables.append(table)
    print(table.name, table.ref)'''

#attempts to use writer method of reading Excel workbook into python; have to look into this further
'''writer = pd.ExcelWriter(fn, engine="openpyxl")
writer.book = wb
sotw_points.to_excel(writer, sheet_name = "sotw_points")
writer.save()
writer.close()'''

#uses read_excel and to_excel to load and save Excel workbooks
'''excel_points = pd.read_excel(path.join(DATA_DIR, 'Song of the Week_{}.xlsx'.format(date_xlsx_fn)),sheet_name="Points Calculation Sheet", index_col=0, usecols=(0,1,2,3,4))
sotw_points.to_excel('Song of the Week_{}.xlsx'.format(date_xlsx_fn),sheet_name="Points Calculation Sheet")'''